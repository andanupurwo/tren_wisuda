import datetime
import glob
import os
import re

from openpyxl import load_workbook
from typing import Optional

try:
    from backend.db import get_conn
except ModuleNotFoundError:
    from db import get_conn


def normalize_label(label: str) -> str:
    return re.sub(r"[^a-z0-9]", "", label.lower().strip())


COLUMN_MAP = {
    "fakultas": "fakultas",
    "prodi": "prodi",
    "program": "program",
    "stausawal": "status_awal",
    "statusawal": "status_awal",
    "stausawalmhs": "status_awal",
    "statusawalmhs": "status_awal",
    "mhsangkatan": "mhs_angkatan",
    "angkatan": "mhs_angkatan",
    "pesertavalid": "peserta_valid",
    "npm": "npm",
    "nama": "nama",
    "jeniskelamin": "jenis_kelamin",
    "ukurantoga": "ukuran_toga",
    "catatan": "catatan",
    "email": "email",
    "telepon1": "telepon1",
    "telepon2": "telepon2",
    "tempatlahir": "tempat_lahir",
    "tanggallahir": "tanggal_lahir",
    "tanggallulus": "tanggal_lulus",
    "masastudibulan": "masa_studi_bulan",
    "masastuditahun": "masa_studi_tahun",
    "namaayah": "nama_ayah",
    "pekerjaanortu": "pekerjaan_ortu",
    "jabatanortu": "jabatan_ortu",
    "ipk": "ipk",
    "sks": "sks",
    "predikat": "predikat",
    "judultaskripsi": "judul_ta_skripsi",
    "catatanupt": "catatan_upt",
    "catatanrc": "catatan_rc",
    "catatandpk": "catatan_dpk",
    "catatanbpc": "catatan_bpc",
    "catatandaak": "catatan_daak",
    "approveupt": "approve_upt",
    "approverc": "approve_rc",
    "approvedpk": "approve_dpk",
    "approvebpc": "approve_bpc",
    "approvedaak": "approve_daak",
}

DB_COLUMNS = [
    "npm",
    "periode",
    "fakultas",
    "prodi",
    "program",
    "status_awal",
    "mhs_angkatan",
    "peserta_valid",
    "nama",
    "jenis_kelamin",
    "ukuran_toga",
    "catatan",
    "email",
    "telepon1",
    "telepon2",
    "tempat_lahir",
    "tanggal_lahir",
    "tanggal_lulus",
    "masa_studi_bulan",
    "masa_studi_tahun",
    "nama_ayah",
    "pekerjaan_ortu",
    "jabatan_ortu",
    "ipk",
    "sks",
    "predikat",
    "judul_ta_skripsi",
    "catatan_upt",
    "catatan_rc",
    "catatan_dpk",
    "catatan_bpc",
    "catatan_daak",
    "approve_upt",
    "approve_rc",
    "approve_dpk",
    "approve_bpc",
    "approve_daak",
]

BOOL_COLUMNS = {
    "peserta_valid",
    "approve_upt",
    "approve_rc",
    "approve_dpk",
    "approve_bpc",
    "approve_daak",
}

DATE_COLUMNS = {"tanggal_lahir", "tanggal_lulus"}
INT_COLUMNS = {"masa_studi_bulan", "sks"}
FLOAT_COLUMNS = {"masa_studi_tahun", "ipk"}

def to_text(value):
    if value is None:
        return None
    if hasattr(value, "strftime"):
        try:
            return value.strftime("%Y-%m-%d")
        except Exception:
            return str(value)
    return str(value)


def to_bool(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0

    text = str(value).strip().lower()
    if text in {"", "-", "n/a", "na"}:
        return None
    if text in {"valid"}:
        return True
    if text in {"tidak valid", "tdk valid", "invalid"}:
        return False
    if text in {"y", "yes", "ya", "true", "1", "t", "v", "ok", "✓"}:
        return True
    if text in {"n", "no", "tidak", "false", "0", "f", "x"}:
        return False
    return None


def to_date(value):
    if value is None:
        return None
    if hasattr(value, "date"):
        try:
            return value.date()
        except Exception:
            return None
    if isinstance(value, (int, float)):
        try:
            return (datetime.datetime(1899, 12, 30) + datetime.timedelta(
                days=int(value)
            )).date()
        except Exception:
            return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.datetime.strptime(text, fmt).date()
            except Exception:
                continue
    return None


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def parse_periode(filename: str) -> Optional[int]:
    match = re.search(r"periode\s*(\d+)", filename.lower())
    if not match:
        return None
    return int(match.group(1))


def load_file(path: str):
    periode = parse_periode(os.path.basename(path))
    if periode is None:
        raise ValueError(f"Periode tidak ditemukan dari nama file: {path}")

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    header_cells = list(ws[1])
    headers = [normalize_label(str(cell.value or "")) for cell in header_cells]
    col_map = {
        idx: COLUMN_MAP[header]
        for idx, header in enumerate(headers)
        if header in COLUMN_MAP
    }
    unknown_headers = [
        header_cells[idx].value
        for idx, header in enumerate(headers)
        if header and header not in COLUMN_MAP
    ]
    if unknown_headers:
        print(f"Kolom tidak dikenal di {os.path.basename(path)}: {unknown_headers}")

    if "npm" not in col_map.values():
        raise ValueError("Kolom wajib tidak ditemukan: ['npm']")

    rows_raw = []
    rows_norm = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record_raw = {col: None for col in DB_COLUMNS}
        record_norm = {col: None for col in DB_COLUMNS}
        record_raw["periode"] = periode
        record_norm["periode"] = periode

        for idx, value in enumerate(row):
            if idx not in col_map:
                continue
            col_name = col_map[idx]
            if col_name == "npm":
                record_raw[col_name] = (
                    str(value).strip() if value is not None else None
                )
                record_norm[col_name] = record_raw[col_name]
                continue
            record_raw[col_name] = to_text(value)
            if col_name in BOOL_COLUMNS:
                record_norm[col_name] = to_bool(value)
            elif col_name in DATE_COLUMNS:
                record_norm[col_name] = to_date(value)
            elif col_name in INT_COLUMNS:
                number = to_number(value)
                record_norm[col_name] = int(number) if number is not None else None
            elif col_name in FLOAT_COLUMNS:
                record_norm[col_name] = to_number(value)
            else:
                record_norm[col_name] = to_text(value)

        if not record_raw["npm"]:
            continue
        rows_raw.append([record_raw[col] for col in DB_COLUMNS])
        rows_norm.append([record_norm[col] for col in DB_COLUMNS])
    if not rows_raw:
        return 0

    insert_cols = ", ".join(DB_COLUMNS)
    update_cols = [col for col in DB_COLUMNS if col not in {"npm", "periode"}]
    set_clause = ", ".join([f"{col} = EXCLUDED.{col}" for col in update_cols])
    set_clause += ", updated_at = NOW()"

    placeholders = ", ".join(["%s"] * len(DB_COLUMNS))
    query_raw = (
        f"INSERT INTO peserta_wisuda_raw ({insert_cols}) VALUES ({placeholders}) "
        f"ON CONFLICT (npm, periode) DO UPDATE SET {set_clause}"
    )
    query_norm = (
        f"INSERT INTO peserta_wisuda ({insert_cols}) VALUES ({placeholders}) "
        f"ON CONFLICT (npm, periode) DO UPDATE SET {set_clause}"
    )

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.executemany(query_raw, rows_raw)
            cur.executemany(query_norm, rows_norm)
        conn.commit()

    return len(rows_raw)


def main():
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    data_dir = os.path.join(base_dir, "history_peserta_wisuda")
    
    # Fallback for Docker environment (where script is in /app and data is in /app/history_peserta_wisuda)
    if not os.path.exists(data_dir):
        data_dir = os.path.join(os.path.dirname(__file__), "history_peserta_wisuda")

    pattern = os.path.join(data_dir, "*.xlsx")
    print(f"DEBUG: base_dir={base_dir}")
    print(f"DEBUG: data_dir={data_dir}")
    print(f"DEBUG: pattern={pattern}")
    files = sorted(glob.glob(pattern))
    print(f"DEBUG: files found={files}")

    if not files:
        print("Tidak ada file .xlsx ditemukan.")
        return

    total_rows = 0
    for path in files:
        count = load_file(path)
        total_rows += count
        print(f"{os.path.basename(path)} -> {count} baris")

    print(f"Selesai. Total baris: {total_rows}")


if __name__ == "__main__":
    main()
